"""
Build a styled gap-analysis Excel workbook (results tables + quadrant charts).

Charts are real Excel chart objects bound to Z_X / Z_Y / Label ranges (not pictures).
Style matches the reference biplot: section title, quadrant marker colors, statement
labels with leader lines, Performance (z) / Importance (z) titles, centre crosshair
and plot border, no tick numbers, no legend.
"""

from __future__ import annotations

import re
import zipfile
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.chart.label import DataLabel, DataLabelList
from openpyxl.chart.marker import Marker
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.drawing.spreadsheet_drawing import AbsoluteAnchor
from openpyxl.drawing.xdr import XDRPoint2D, XDRPositiveSize2D
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.units import cm_to_EMU, pixels_to_EMU
from openpyxl.worksheet.worksheet import Worksheet

# Theme: white + space blue (matches app --space-blue)
BLUE = "1B254A"
BLUE_LIGHT = "E8ECF4"
BLUE_MID = "C5CEE0"
WHITE = "FFFFFF"
TEXT = "1B254A"
MUTED = "5A6478"
CHART_MARKER_BLUE = "2E75B6"

QUADRANT_COLORS = {
    "urgent": "990000",
    "maintain": "38761D",
    "low": "BF9000",
    "overkill": "1155CC",
}

QUADRANT_LABELS = {
    "urgent": "Low Performance High Importance",
    "maintain": "High Performance High Importance",
    "low": "Low Performance Low Importance",
    "overkill": "High Performance Low Importance",
}

TABLE_HEADERS = [
    "Statement",
    "Section",
    "Performance",
    "Importance",
    "Reduced importance",
    "Z performance. (X)",
    "Z importance. (Y)",
    "Position",
]

# Chart source: K=Z_X, L=Z_Y, M=Quadrant, N=Label
CHART_Z_COL = 11
# I1 = chart title (section name), J1 = chart anchor row
CHART_TITLE_CELL = (1, 9)  # I1
CHART_ANCHOR_CELL = (1, 10)  # J1

# Populated per export request and passed into chart XML post-processing.
CHART_CROSSHAIR_SERIES = 2
BI_PLOT_AXIS_GREY = "C9CED8"
CHART_OUTER_BORDER = "6B7280"
# Crosshair stops short of plot edge (fraction of axis span per side).
CROSSHAIR_INSET_FRAC = 0.05


def _escape_xml_text(text: str) -> str:
    return (
        str(text or "")
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
    )


def _label_cell_formula(sheet_title: str, row: int) -> str:
    col = get_column_letter(CHART_Z_COL + 3)
    safe = str(sheet_title or "Sheet").replace("'", "''")
    return f"'{safe}'!${col}${row}"


def _label_tx_xml(formula: str, text: str) -> str:
    return (
        "<tx><strRef>"
        f"<f>{_escape_xml_text(formula)}</f>"
        "<strCache>"
        '<ptCount val="1"/>'
        '<pt idx="0">'
        f"<v>{_escape_xml_text(text)}</v>"
        "</pt></strCache>"
        "</strRef></tx>"
    )


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold: bool = False, color: str = TEXT, size: int = 11) -> Font:
    return Font(name="Calibri", bold=bold, color=color, size=size)


def _thin_border() -> Border:
    side = Side(style="thin", color=BLUE_MID)
    return Border(left=side, right=side, top=side, bottom=side)


def _safe_sheet_title(name: str, used: set[str]) -> str:
    cleaned = re.sub(r'[\\/*?:\[\]]', "-", str(name or "Sheet")).strip() or "Sheet"
    cleaned = cleaned[:31]
    base = cleaned
    n = 2
    while cleaned in used:
        suffix = f" ({n})"
        cleaned = (base[: 31 - len(suffix)] + suffix)[:31]
        n += 1
    used.add(cleaned)
    return cleaned


def _position_label(quadrant: str, z_perf: float | None, z_imp: float | None) -> str:
    if quadrant in QUADRANT_LABELS:
        return QUADRANT_LABELS[quadrant]
    zp = float(z_perf or 0)
    zi = float(z_imp or 0)
    if zp >= 0 and zi >= 0:
        return QUADRANT_LABELS["maintain"]
    if zp < 0 and zi < 0:
        return QUADRANT_LABELS["low"]
    if zp < 0 and zi >= 0:
        return QUADRANT_LABELS["urgent"]
    return QUADRANT_LABELS["overkill"]


def _pct(val: Any) -> str | float:
    if val is None or val == "":
        return "-"
    try:
        return float(val)
    except (TypeError, ValueError):
        return "-"


def _z(val: Any) -> str | float:
    if val is None or val == "":
        return "-"
    try:
        return round(float(val), 3)
    except (TypeError, ValueError):
        return "-"


def _style_header_row(ws: Worksheet, row: int, col_count: int) -> None:
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = _fill(BLUE)
        cell.font = _font(bold=True, color=WHITE, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _thin_border()


def _style_data_cell(cell, *, bold: bool = False, color: str | None = None, center: bool = False) -> None:
    cell.font = _font(bold=bold, color=color or TEXT)
    cell.alignment = Alignment(
        horizontal="center" if center else "left",
        vertical="center",
        wrap_text=True,
    )
    cell.border = _thin_border()


def _write_title(ws: Worksheet, title: str, subtitle: str | None = None) -> int:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = _font(bold=True, color=BLUE, size=16)
    title_cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 24

    if subtitle:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)
        sub = ws.cell(row=2, column=1, value=subtitle)
        sub.font = _font(color=MUTED, size=10)
        return 3
    return 2


def _write_table(
    ws: Worksheet,
    start_row: int,
    *,
    overall_csat: Any,
    sections: list[dict[str, Any]],
    include_section_headers: bool,
) -> tuple[int, list[dict[str, Any]]]:
    """Write results table. Returns (next_row, flat statement rows)."""
    header_row = start_row
    for i, h in enumerate(TABLE_HEADERS, start=1):
        ws.cell(row=header_row, column=i, value=h)
    _style_header_row(ws, header_row, len(TABLE_HEADERS))
    ws.row_dimensions[header_row].height = 30

    row = header_row + 1
    flat: list[dict[str, Any]] = []

    values = [
        "Overall satisfaction",
        "-",
        _pct(overall_csat),
        "-",
        "-",
        "-",
        "-",
        "-",
    ]
    for col, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col, value=val)
        _style_data_cell(cell, bold=True, color="990000" if col in (1, 3) else TEXT, center=col > 1)
        cell.fill = _fill(BLUE_LIGHT)
    if isinstance(values[2], float):
        ws.cell(row=row, column=3).number_format = "0.0"
    row += 1

    for sec in sections:
        sec_name = sec.get("section_name") or f"Section {sec.get('section')}"
        stmts = sec.get("statements") or []

        if include_section_headers:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
            cell = ws.cell(row=row, column=1, value=sec_name)
            cell.fill = _fill(BLUE_MID)
            cell.font = _font(bold=True, color=BLUE)
            cell.alignment = Alignment(vertical="center")
            for col in range(1, 9):
                ws.cell(row=row, column=col).fill = _fill(BLUE_MID)
                ws.cell(row=row, column=col).border = _thin_border()
            row += 1

        for stmt in stmts:
            quadrant = stmt.get("quadrant") or ""
            pos = _position_label(
                quadrant, stmt.get("z_performance"), stmt.get("z_importance")
            )
            # Swapped display: Importance <- section-reduced; Reduced importance <- global
            row_vals = [
                stmt.get("label") or stmt.get("column") or "",
                sec_name,
                _pct(stmt.get("performance")),
                _pct(stmt.get("importance_section")),
                _pct(stmt.get("importance")),
                _z(stmt.get("z_performance")),
                _z(stmt.get("z_importance")),
                pos,
            ]

            for col, val in enumerate(row_vals, start=1):
                cell = ws.cell(row=row, column=col, value=val)
                color = QUADRANT_COLORS.get(quadrant, TEXT) if col == 8 else TEXT
                _style_data_cell(cell, color=color, center=col > 1, bold=(col == 8))
                if row % 2 == 0:
                    cell.fill = _fill("F7F8FB")
            for col in (3, 4, 5):
                if isinstance(row_vals[col - 1], float):
                    ws.cell(row=row, column=col).number_format = "0.0"

            flat.append(
                {
                    "label": row_vals[0],
                    "z_performance": stmt.get("z_performance"),
                    "z_importance": stmt.get("z_importance"),
                    "quadrant": quadrant,
                }
            )
            row += 1

    widths = [42, 16, 12, 12, 16, 16, 16, 32]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    return row, flat


def _write_chart_source(
    ws: Worksheet, start_row: int, points: list[dict[str, Any]]
) -> tuple[int, int, int] | None:
    """
    Write Z_X / Z_Y / Label ranges used by the chart + VBA.
    Returns (header_row, first_data_row, last_data_row) or None.
    """
    if not points:
        return None

    note_row = start_row
    note = ws.cell(
        row=note_row,
        column=CHART_Z_COL,
        value=(
            "Chart data source -> X axis reads Z_X (Performance z), "
            "Y axis reads Z_Y (Importance z), point labels read Label. "
            "Quadrant colors read Quadrant."
        ),
    )
    note.font = _font(bold=True, color=BLUE, size=10)
    note.alignment = Alignment(wrap_text=True, vertical="center")
    ws.merge_cells(
        start_row=note_row,
        start_column=CHART_Z_COL,
        end_row=note_row,
        end_column=CHART_Z_COL + 3,
    )
    ws.row_dimensions[note_row].height = 32

    header_row = note_row + 1
    headers = (
        (CHART_Z_COL, "Z_X"),
        (CHART_Z_COL + 1, "Z_Y"),
        (CHART_Z_COL + 2, "Quadrant"),
        (CHART_Z_COL + 3, "Label"),
    )
    for col, text in headers:
        cell = ws.cell(row=header_row, column=col, value=text)
        cell.font = _font(bold=True, color=WHITE, size=10)
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Sub-headers so it's obvious what each z column is
    sub_row = header_row + 1
    sub_headers = (
        (CHART_Z_COL, "Performance z -> chart X"),
        (CHART_Z_COL + 1, "Importance z -> chart Y"),
        (CHART_Z_COL + 2, "Marker color"),
        (CHART_Z_COL + 3, "Data label + leader line"),
    )
    for col, text in sub_headers:
        cell = ws.cell(row=sub_row, column=col, value=text)
        cell.font = _font(color=MUTED, size=8)
        cell.fill = PatternFill("solid", fgColor=BLUE_LIGHT)

    first = sub_row + 1
    r = first
    for p in points:
        try:
            x_val = float(p.get("z_performance"))
            y_val = float(p.get("z_importance"))
        except (TypeError, ValueError):
            continue
        ws.cell(row=r, column=CHART_Z_COL, value=x_val)
        ws.cell(row=r, column=CHART_Z_COL + 1, value=y_val)
        ws.cell(row=r, column=CHART_Z_COL + 2, value=p.get("quadrant") or "")
        ws.cell(row=r, column=CHART_Z_COL + 3, value=p.get("label") or "")
        r += 1

    if r == first:
        return None

    last = r - 1
    # Explicit range note under the block
    range_note = ws.cell(
        row=last + 1,
        column=CHART_Z_COL,
        value=(
            f"Graph series bound to "
            f"{get_column_letter(CHART_Z_COL)}{first}:{get_column_letter(CHART_Z_COL)}{last} "
            f"(Z_X) and "
            f"{get_column_letter(CHART_Z_COL + 1)}{first}:{get_column_letter(CHART_Z_COL + 1)}{last} "
            f"(Z_Y); labels from "
            f"{get_column_letter(CHART_Z_COL + 3)}{first}:{get_column_letter(CHART_Z_COL + 3)}{last}."
        ),
    )
    range_note.font = _font(color=MUTED, size=8)
    ws.merge_cells(
        start_row=last + 1,
        start_column=CHART_Z_COL,
        end_row=last + 1,
        end_column=CHART_Z_COL + 3,
    )

    for col, width in (
        (CHART_Z_COL, 14),
        (CHART_Z_COL + 1, 14),
        (CHART_Z_COL + 2, 12),
        (CHART_Z_COL + 3, 36),
    ):
        ws.column_dimensions[get_column_letter(col)].width = width

    return header_row, first, last


def _table_width_px(ws: Worksheet, last_col: int = 8) -> float:
    """Approximate pixel width of columns A..last_col (Excel char-width heuristic)."""
    total_chars = 0.0
    for col in range(1, last_col + 1):
        letter = get_column_letter(col)
        width = ws.column_dimensions[letter].width
        total_chars += float(width) if width else 8.43
    return total_chars * 7.0


def _row_top_px(ws: Worksheet, anchor_row: int) -> float:
    """Approximate pixel offset from top of sheet to anchor_row."""
    top = 0.0
    for r in range(1, max(1, anchor_row)):
        height = ws.row_dimensions[r].height
        top += float(height) if height else 15.0
    return top


def _equal_axis_extent(
    xs: list[float],
    ys: list[float],
    *,
    pad: float = 0.45,
    min_half: float = 2.5,
) -> tuple[float, float]:
    """
    Same numeric range for X and Y (symmetric about 0), matching the app biplot.
    Half-extent = max(|z| + pad, min_half) across both axes.
    """
    extent = float(min_half)
    for v in list(xs) + list(ys):
        try:
            extent = max(extent, abs(float(v)) + pad)
        except (TypeError, ValueError):
            continue
    extent = round(extent, 3)
    return -extent, extent


def _style_biplot_axes(chart: ScatterChart) -> None:
    """Scale/titles only — centre crosshair + outer box drawn as line series."""
    axis_hidden = GraphicalProperties(ln=LineProperties(noFill=True))
    chart.x_axis.axPos = "b"
    chart.y_axis.axPos = "l"
    for axis in (chart.x_axis, chart.y_axis):
        axis.spPr = axis_hidden
        axis.majorTickMark = "none"
        axis.minorTickMark = "none"
        axis.tickLblPos = "none"
        axis.crosses = "autoZero"
        axis.majorGridlines = None
        axis.minorGridlines = None


def _crosshair_line_series(
    ws: Worksheet, row_a: int, row_b: int, *, title: str
) -> Series:
    xvalues = Reference(ws, min_col=CHART_Z_COL, min_row=row_a, max_row=row_b)
    yvalues = Reference(ws, min_col=CHART_Z_COL + 1, min_row=row_a, max_row=row_b)
    series = Series(yvalues, xvalues, title=title)
    series.graphicalProperties = GraphicalProperties(
        ln=LineProperties(solidFill=BI_PLOT_AXIS_GREY, w=9525)
    )
    series.marker = Marker(symbol="none")
    return series


def _crosshair_inner_bounds(axis_min: float, axis_max: float) -> tuple[float, float]:
    span = axis_max - axis_min
    inset = span * CROSSHAIR_INSET_FRAC
    return axis_min + inset, axis_max - inset


def _write_crosshair_source(
    ws: Worksheet, last_row: int, axis_min: float, axis_max: float
) -> tuple[tuple[int, int], tuple[int, int]]:
    """Scratch Z rows for inset centre horizontal + vertical crosshair lines."""
    inner_min, inner_max = _crosshair_inner_bounds(axis_min, axis_max)
    base = last_row + 2  # last_row+1 is merged range note
    h1, h2 = base, base + 1
    v1, v2 = base + 2, base + 3
    ws.cell(row=h1, column=CHART_Z_COL, value=inner_min)
    ws.cell(row=h1, column=CHART_Z_COL + 1, value=0)
    ws.cell(row=h2, column=CHART_Z_COL, value=inner_max)
    ws.cell(row=h2, column=CHART_Z_COL + 1, value=0)
    ws.cell(row=v1, column=CHART_Z_COL, value=0)
    ws.cell(row=v1, column=CHART_Z_COL + 1, value=inner_min)
    ws.cell(row=v2, column=CHART_Z_COL, value=0)
    ws.cell(row=v2, column=CHART_Z_COL + 1, value=inner_max)
    return (h1, h2), (v1, v2)


def _quadrant_marker_color(quadrant: str) -> str:
    return QUADRANT_COLORS.get((quadrant or "").strip().lower(), CHART_MARKER_BLUE)


def _add_openpyxl_chart(
    ws: Worksheet,
    *,
    chart_title: str,
    first: int,
    last: int,
    anchor_row: int,
) -> list[tuple[str, str]]:
    """
    Biplot-style scatter: equal X/Y axis dimensions (shared z extent),
    centered under the table, markers only, labels from Label column,
    marker colors by quadrant.
    """
    xs: list[float] = []
    ys: list[float] = []
    for row in range(first, last + 1):
        try:
            xs.append(float(ws.cell(row=row, column=CHART_Z_COL).value))
            ys.append(float(ws.cell(row=row, column=CHART_Z_COL + 1).value))
        except (TypeError, ValueError):
            continue

    axis_min, axis_max = _equal_axis_extent(xs, ys)

    chart = ScatterChart()
    chart.title = chart_title
    # lineMarker so crosshair/box line series render; data series stay marker-only
    chart.scatterStyle = "lineMarker"
    chart.x_axis.title = "Performance (z)"
    chart.y_axis.title = "Importance (z)"
    # Same +/- extent on both axes so (0,0) is centered (not shifted)
    chart.x_axis.scaling.min = axis_min
    chart.x_axis.scaling.max = axis_max
    chart.y_axis.scaling.min = axis_min
    chart.y_axis.scaling.max = axis_max
    chart.x_axis.crosses = "autoZero"
    chart.y_axis.crosses = "autoZero"
    span = axis_max - axis_min
    major = round(span / 10.0, 2) if span > 0 else 0.5
    chart.x_axis.majorUnit = major
    chart.y_axis.majorUnit = major
    _style_biplot_axes(chart)
    chart.legend = None
    h_rows, v_rows = _write_crosshair_source(ws, last, axis_min, axis_max)
    chart.series.append(_crosshair_line_series(ws, h_rows[0], h_rows[1], title="_cross_h"))
    chart.series.append(_crosshair_line_series(ws, v_rows[0], v_rows[1], title="_cross_v"))
    # Square chart object => equal visual quadrant sizes
    chart_size_cm = 16.0
    chart.height = chart_size_cm
    chart.width = chart_size_cm

    label_batch: list[tuple[str, str]] = []

    for row in range(first, last + 1):
        xvalues = Reference(ws, min_col=CHART_Z_COL, min_row=row, max_row=row)
        yvalues = Reference(ws, min_col=CHART_Z_COL + 1, min_row=row, max_row=row)
        label = ws.cell(row=row, column=CHART_Z_COL + 3).value
        quadrant = str(ws.cell(row=row, column=CHART_Z_COL + 2).value or "")
        title = str(label).strip() if label not in (None, "") else f"Point {row}"
        if len(title) > 200:
            title = title[:197] + "..."
        label_batch.append((_label_cell_formula(ws.title, row), title))
        series = Series(yvalues, xvalues, title=title)
        # Markers only — no connecting line between points
        series.graphicalProperties = GraphicalProperties(
            ln=LineProperties(noFill=True)
        )
        series.marker = Marker(symbol="circle", size=8)
        series.marker.graphicalProperties = GraphicalProperties(
            solidFill=_quadrant_marker_color(quadrant),
            ln=LineProperties(noFill=True),
        )
        dLbls = DataLabelList()
        dLbls.showSerName = False
        dLbls.showVal = False
        dLbls.showCatName = False
        dLbls.showPercent = False
        dLbls.showLegendKey = False
        dLbls.showLeaderLines = True
        dLbls.dLblPos = "r"
        # Per-point label entry (layout offset added in _patch_chart_leader_lines)
        point_lbl = DataLabel(idx=0)
        point_lbl.showSerName = False
        point_lbl.showVal = False
        point_lbl.showCatName = False
        point_lbl.showLegendKey = False
        point_lbl.showLeaderLines = True
        point_lbl.dLblPos = "r"
        dLbls.dLbl.append(point_lbl)
        series.dLbls = dLbls
        chart.series.append(series)
    # Center under results table (A-H)
    table_px = _table_width_px(ws, last_col=8)
    chart_px = chart_size_cm * (96.0 / 2.54)
    left_px = max(0.0, (table_px - chart_px) / 2.0)
    top_px = _row_top_px(ws, anchor_row)

    chart.anchor = AbsoluteAnchor(
        pos=XDRPoint2D(x=pixels_to_EMU(left_px), y=pixels_to_EMU(top_px)),
        ext=XDRPositiveSize2D(cx=cm_to_EMU(chart_size_cm), cy=cm_to_EMU(chart_size_cm)),
    )
    ws.add_chart(chart)
    return label_batch


def _inject_dLbl_layout(dLbl_xml: str, ox: float, oy: float) -> str:
    """Insert manualLayout into a <dLbl>…</dLbl> block if missing."""
    if "<layout>" in dLbl_xml or "<c:layout>" in dLbl_xml:
        return dLbl_xml
    layout = (
        f"<layout><manualLayout>"
        f'<xMode val="factor"/><yMode val="factor"/>'
        f'<x val="{ox:.4f}"/><y val="{oy:.4f}"/>'
        f"</manualLayout></layout>"
    )
    # Place layout right after <dLbl>…<idx …/>
    return re.sub(
        r"(<dLbl>\s*<idx[^/]*/>)",
        r"\1" + layout,
        dLbl_xml,
        count=1,
        flags=re.I,
    )


def _patch_chart_xml_leader_lines(
    xml: str, label_cells: list[tuple[str, str]]
) -> str:
    """
    Excel draws scatter leader lines only when labels are manually offset.
    Inject factor-based layout offsets and cell-based label text (no legend key).
    """
    offsets = (
        (0.05, -0.06),
        (0.05, 0.045),
        (-0.14, -0.06),
        (-0.14, 0.045),
    )
    ser_idx = 0

    def patch_ser(match: re.Match[str]) -> str:
        nonlocal ser_idx
        block = match.group(0)
        if ser_idx < CHART_CROSSHAIR_SERIES:
            ser_idx += 1
            return block
        ox, oy = offsets[(ser_idx - CHART_CROSSHAIR_SERIES) % 4]
        formula, text = ("", "")
        label_idx = ser_idx - CHART_CROSSHAIR_SERIES
        if label_idx < len(label_cells):
            formula, text = label_cells[label_idx]
        tx_xml = _label_tx_xml(formula, text) if formula else ""
        ser_idx += 1

        def patch_dlbls(m: re.Match[str]) -> str:
            dlbls = m.group(0)
            # Ensure showLeaderLines is on
            if "showLeaderLines" not in dlbls:
                dlbls = dlbls.replace("</dLbls>", '<showLeaderLines val="1"/></dLbls>')
            else:
                dlbls = re.sub(
                    r'<showLeaderLines[^/]*/>',
                    '<showLeaderLines val="1"/>',
                    dlbls,
                )
            dlbls = re.sub(
                r"<showSerName[^/]*/>",
                '<showSerName val="0"/>',
                dlbls,
            )
            dlbls = re.sub(
                r"<showLegendKey[^/]*/>",
                '<showLegendKey val="0"/>',
                dlbls,
            )
            if "showSerName" not in dlbls:
                dlbls = dlbls.replace("</dLbls>", '<showSerName val="0"/></dLbls>')
            if "showLegendKey" not in dlbls:
                dlbls = dlbls.replace("</dLbls>", '<showLegendKey val="0"/></dLbls>')
            # Ensure position is not center
            if "dLblPos" not in dlbls:
                dlbls = dlbls.replace("</dLbls>", '<dLblPos val="r"/></dLbls>')

            if re.search(r"<dLbl[\s>]", dlbls):
                def patch_point_lbl(mm: re.Match[str]) -> str:
                    dlbl = mm.group(0)
                    if tx_xml and "<tx>" not in dlbl:
                        dlbl = re.sub(
                            r"(<dLbl>\s*<idx[^/]*/>)",
                            r"\1" + tx_xml,
                            dlbl,
                            count=1,
                            flags=re.I,
                        )
                    dlbl = re.sub(
                        r"<showSerName[^/]*/>",
                        '<showSerName val="0"/>',
                        dlbl,
                    )
                    dlbl = re.sub(
                        r"<showLegendKey[^/]*/>",
                        '<showLegendKey val="0"/>',
                        dlbl,
                    )
                    if "showSerName" not in dlbl:
                        dlbl = dlbl.replace("</dLbl>", '<showSerName val="0"/></dLbl>')
                    if "showLegendKey" not in dlbl:
                        dlbl = dlbl.replace(
                            "</dLbl>", '<showLegendKey val="0"/></dLbl>'
                        )
                    return _inject_dLbl_layout(dlbl, ox, oy)

                dlbls = re.sub(
                    r"<dLbl>.*?</dLbl>",
                    patch_point_lbl,
                    dlbls,
                    flags=re.S,
                )
            else:
                injected = (
                    f'<dLbl><idx val="0"/>'
                    f"{tx_xml}"
                    f"<layout><manualLayout>"
                    f'<xMode val="factor"/><yMode val="factor"/>'
                    f'<x val="{ox:.4f}"/><y val="{oy:.4f}"/>'
                    f"</manualLayout></layout>"
                    f'<dLblPos val="r"/>'
                    f'<showSerName val="0"/><showVal val="0"/>'
                    f'<showLegendKey val="0"/>'
                    f'<showLeaderLines val="1"/>'
                    f"</dLbl>"
                )
                dlbls = dlbls.replace("<dLbls>", f"<dLbls>{injected}", 1)
            return dlbls

        return re.sub(r"<dLbls>.*?</dLbls>", patch_dlbls, block, flags=re.S)

    return re.sub(r"<ser>.*?</ser>", patch_ser, xml, flags=re.S)


def _patch_chart_xml_restore_axes(xml: str) -> str:
    """Hide edge axes; white plot; no tick numbers; no gridlines."""

    axis_hidden = (
        '<spPr><a:ln xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">'
        "<a:noFill/></a:ln></spPr>"
    )
    plot_fill = (
        '<spPr xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">'
        '<a:solidFill><a:srgbClr val="FFFFFF"/></a:solidFill>'
        "<a:ln><a:noFill/></a:ln></spPr>"
    )

    def patch_axis(match: re.Match[str]) -> str:
        block = match.group(0)
        ns = match.group(1) or ""
        tag = match.group(2)
        open_tag = f"<{ns}{tag}>"
        close_tag = f"</{ns}{tag}>"
        if "Performance" in block:
            block = re.sub(
                r"<(?:c:)?axPos[^/]*/>",
                '<axPos val="b"/>',
                block,
            )
            if "axPos" not in block:
                block = block.replace(
                    "<scaling>", '<axPos val="b"/><scaling>', 1
                )
        elif "Importance" in block:
            block = re.sub(
                r"<(?:c:)?axPos[^/]*/>",
                '<axPos val="l"/>',
                block,
            )
            if "axPos" not in block:
                block = block.replace(
                    "<scaling>", '<axPos val="l"/><scaling>', 1
                )
        block = re.sub(
            r"<(?:c:)?tickLblPos[^/]*/>",
            '<tickLblPos val="none"/>',
            block,
        )
        if "tickLblPos" not in block:
            block = block.replace(close_tag, '<tickLblPos val="none"/>' + close_tag, 1)
        block = re.sub(
            r"<(?:c:)?majorTickMark[^/]*/>",
            '<majorTickMark val="none"/>',
            block,
        )
        if "majorTickMark" not in block:
            block = block.replace(
                close_tag, '<majorTickMark val="none"/>' + close_tag, 1
            )
        block = re.sub(r"<crosses[^/]*/>", '<crosses val="autoZero"/>', block)
        block = re.sub(r"<crossesAt[^/]*/>", "", block)
        block = re.sub(r"<spPr>.*?</spPr>", axis_hidden, block, count=1, flags=re.S)
        if "spPr" not in block:
            block = block.replace(open_tag, open_tag + axis_hidden, 1)
        block = re.sub(r"<majorGridlines>.*?</majorGridlines>", "", block, flags=re.S)
        block = re.sub(r"<minorGridlines>.*?</minorGridlines>", "", block, flags=re.S)
        block = re.sub(r"<majorGridlines[^/]*/>", "", block)
        block = re.sub(r"<minorGridlines[^/]*/>", "", block)
        return block

    xml = re.sub(
        r"<(?:(c:)?)(valAx|catAx)>.*?</(?:(c:)?)(?:valAx|catAx)>",
        patch_axis,
        xml,
        flags=re.S,
    )
    xml = re.sub(
        r"</valAx>\s*<spPr><a:ln[^>]*>.*?</a:ln></spPr>\s*(?=</plotArea>)",
        "</valAx>",
        xml,
        flags=re.S,
    )
    if not re.search(r"<plotArea>\s*<spPr", xml):
        xml = xml.replace("<plotArea>", f"<plotArea>{plot_fill}", 1)

    plot_layout = (
        "<layout><manualLayout>"
        '<layoutTarget val="inner"/>'
        '<xMode val="edge"/><yMode val="edge"/>'
        '<x val="0.05"/><y val="0.06"/>'
        '<w val="0.90"/><h val="0.84"/>'
        "</manualLayout></layout>"
    )
    if "<layout>" not in xml.split("</plotArea>", 1)[0]:
        xml = xml.replace("</plotArea>", plot_layout + "</plotArea>", 1)
    xml = re.sub(
        r"<scatterStyle[^/]*/>",
        '<scatterStyle val="lineMarker"/>',
        xml,
    )
    if "scatterStyle" not in xml:
        xml = xml.replace(
            "<scatterChart>",
            '<scatterChart><scatterStyle val="lineMarker"/>',
            1,
        )

    def patch_crosshair_ser(match: re.Match[str]) -> str:
        block = match.group(0)
        if "_cross_" not in block:
            return block
        block = re.sub(r"<smooth[^/]*/>", '<smooth val="0"/>', block)
        if "smooth" not in block:
            block = block.replace("</ser>", '<smooth val="0"/></ser>', 1)
        block = re.sub(r"<marker>.*?</marker>", "<marker><symbol val=\"none\"/></marker>", block, flags=re.S)
        if "<marker" not in block:
            block = block.replace("</ser>", '<marker><symbol val="none"/></marker></ser>', 1)
        block = re.sub(r"<dLbls>.*?</dLbls>", "", block, flags=re.S)
        return block

    xml = re.sub(r"<ser>.*?</ser>", patch_crosshair_ser, xml, flags=re.S)

    # Outer chart frame only — no inner plot-area border.
    chart_area_frame = (
        '<spPr xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">'
        '<a:solidFill><a:srgbClr val="FFFFFF"/></a:solidFill>'
        '<a:ln w="12700" cap="flat" cmpd="sng">'
        f'<a:solidFill><a:srgbClr val="{CHART_OUTER_BORDER}"/></a:solidFill>'
        '<a:prstDash val="solid"/></a:ln></spPr>'
    )
    if re.search(r"<chartSpace[^>]*>\s*<spPr", xml, flags=re.S):
        xml = re.sub(
            r"(<chartSpace[^>]*>)\s*<spPr>.*?</spPr>",
            r"\1" + chart_area_frame,
            xml,
            count=1,
            flags=re.S,
        )
    else:
        xml = re.sub(
            r"(<chartSpace[^>]*>)",
            r"\1" + chart_area_frame,
            xml,
            count=1,
        )
    if "roundedCorners" not in xml:
        xml = xml.replace("<chart>", '<chart><roundedCorners val="1"/>', 1)
    else:
        xml = re.sub(
            r"<roundedCorners[^/]*/>",
            '<roundedCorners val="1"/>',
            xml,
        )
    return xml


def _patch_chart_xml_hide_legend_keys(xml: str) -> str:
    """Ensure data labels never show colored legend-key squares."""

    def patch_dlbls(match: re.Match[str]) -> str:
        block = match.group(0)
        block = re.sub(r"<showLegendKey[^/]*/>", '<showLegendKey val="0"/>', block)
        block = re.sub(r"<showSerName[^/]*/>", '<showSerName val="0"/>', block)
        if "showLegendKey" not in block:
            block = block.replace("</dLbls>", '<showLegendKey val="0"/></dLbls>')
        if "showSerName" not in block:
            block = block.replace("</dLbls>", '<showSerName val="0"/></dLbls>')
        return block

    xml = re.sub(r"<dLbls>.*?</dLbls>", patch_dlbls, xml, flags=re.S)

    def patch_dlbl(match: re.Match[str]) -> str:
        block = match.group(0)
        block = re.sub(r"<showLegendKey[^/]*/>", '<showLegendKey val="0"/>', block)
        block = re.sub(r"<showSerName[^/]*/>", '<showSerName val="0"/>', block)
        if "showLegendKey" not in block:
            block = block.replace("</dLbl>", '<showLegendKey val="0"/></dLbl>')
        if "showSerName" not in block:
            block = block.replace("</dLbl>", '<showSerName val="0"/></dLbl>')
        return block

    return re.sub(r"<dLbl>.*?</dLbl>", patch_dlbl, xml, flags=re.S)


def _patch_chart_xml(xml: str, label_cells: list[tuple[str, str]] | None = None) -> str:
    try:
        xml = _patch_chart_xml_leader_lines(xml, label_cells or [])
        xml = _patch_chart_xml_hide_legend_keys(xml)
    except Exception:
        pass
    try:
        xml = _patch_chart_xml_restore_axes(xml)
    except Exception:
        pass
    return xml


def _patch_workbook_charts(
    data: bytes, label_batches: list[list[tuple[str, str]]]
) -> bytes:
    """Post-process chart XML (labels, leader lines, axes)."""
    src = zipfile.ZipFile(BytesIO(data), "r")
    out = BytesIO()
    chart_idx = 0
    with zipfile.ZipFile(out, "w", compression=zipfile.ZIP_DEFLATED) as dst:
        for info in src.infolist():
            raw = src.read(info.filename)
            name = info.filename
            if name.startswith("xl/charts/chart") and name.endswith(".xml"):
                text = raw.decode("utf-8")
                labels = (
                    label_batches[chart_idx]
                    if chart_idx < len(label_batches)
                    else []
                )
                text = _patch_chart_xml(text, labels)
                raw = text.encode("utf-8")
                chart_idx += 1
            dst.writestr(name, raw, compress_type=zipfile.ZIP_DEFLATED)
    src.close()
    return out.getvalue()


def _build_sheet(
    wb: Workbook,
    *,
    title: str,
    sheet_name: str,
    chart_title: str,
    subtitle: str | None,
    overall_csat: Any,
    sections: list[dict[str, Any]],
    include_section_headers: bool,
    used_names: set[str],
    add_chart: bool,
    label_batches: list[list[tuple[str, str]]],
) -> None:
    safe_name = _safe_sheet_title(sheet_name, used_names)
    ws = wb.create_sheet(title=safe_name)
    next_row = _write_title(ws, title, subtitle)
    table_end, points = _write_table(
        ws,
        next_row,
        overall_csat=overall_csat,
        sections=sections,
        include_section_headers=include_section_headers,
    )

    # Chart title = section / sheet name (hidden metadata cell)
    ws.cell(row=CHART_TITLE_CELL[0], column=CHART_TITLE_CELL[1], value=chart_title)
    ws.cell(row=CHART_TITLE_CELL[0], column=CHART_TITLE_CELL[1]).font = _font(
        color=WHITE, size=1
    )

    source_meta = _write_chart_source(ws, table_end + 1, points)
    anchor_row = table_end + 3
    ws.cell(row=CHART_ANCHOR_CELL[0], column=CHART_ANCHOR_CELL[1], value=anchor_row)
    ws.cell(row=CHART_ANCHOR_CELL[0], column=CHART_ANCHOR_CELL[1]).font = _font(
        color=WHITE, size=1
    )

    if source_meta and add_chart:
        _hdr, first, last = source_meta
        label_batch = _add_openpyxl_chart(
            ws,
            chart_title=chart_title,
            first=first,
            last=last,
            anchor_row=anchor_row,
        )
        label_batches.append(label_batch)

    ws.freeze_panes = f"A{next_row + 1}"
    ws.sheet_view.showGridLines = False


def build_gap_analysis_xlsx(result: dict[str, Any], *, filename: str = "") -> bytes:
    """
    Build workbook bytes:
      - Sheet 1: All Sections (full table + chart)
      - Sheet 2+: one sheet per section (chart titled with section name)
    """
    label_batches: list[list[tuple[str, str]]] = []
    table = result.get("table") or {}
    sections = table.get("sections") or []
    overall = table.get("overall_csat")
    if overall is None:
        overall = (result.get("summary_all") or result.get("summary") or {}).get(
            "overall_csat"
        ) or (result.get("summary") or {}).get("overall_performance")

    metric = result.get("metric_label") or ""
    scale = result.get("scale") or ""
    subtitle_bits = []
    if filename:
        subtitle_bits.append(str(filename))
    if scale:
        subtitle_bits.append(f"Scale {scale}")
    if metric:
        subtitle_bits.append(f"Metric: {metric}")
    subtitle = " | ".join(subtitle_bits) if subtitle_bits else None

    wb = Workbook()
    default = wb.active
    wb.remove(default)
    used: set[str] = set()

    _build_sheet(
        wb,
        title="Gap analysis - All sections",
        sheet_name="All Sections",
        chart_title="All Sections",
        subtitle=subtitle,
        overall_csat=overall,
        sections=sections,
        include_section_headers=True,
        used_names=used,
        add_chart=True,
        label_batches=label_batches,
    )

    for sec in sections:
        sec_name = str(sec.get("section_name") or f"Section {sec.get('section')}")
        stmts = sec.get("statements") or []
        sec_overall = None
        if stmts:
            weights = [float(s.get("importance_section") or 0) for s in stmts]
            total_w = sum(weights) or 0
            if total_w:
                sec_overall = round(
                    sum(float(s.get("performance") or 0) * w for s, w in zip(stmts, weights))
                    / total_w,
                    1,
                )
            else:
                sec_overall = round(
                    sum(float(s.get("performance") or 0) for s in stmts) / len(stmts),
                    1,
                )

        _build_sheet(
            wb,
            title=f"Gap analysis - {sec_name}",
            sheet_name=sec_name,
            chart_title=sec_name,
            subtitle=subtitle,
            overall_csat=sec_overall if sec_overall is not None else overall,
            sections=[sec],
            include_section_headers=False,
            used_names=used,
            add_chart=True,
            label_batches=label_batches,
        )

    buf = BytesIO()
    wb.save(buf)
    return _patch_workbook_charts(buf.getvalue(), label_batches)
