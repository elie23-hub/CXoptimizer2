"""Excel export for simulation (bottom-up / top-down) and summary pages."""

from __future__ import annotations

import re
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

BLUE = "1B254A"
BLUE_LIGHT = "E8ECF4"
BLUE_MID = "C5CEE0"
WHITE = "FFFFFF"
TEXT = "1B254A"
MUTED = "5A6478"

QUADRANT_LABELS = {
    "urgent": "Low performance high importance",
    "maintain": "High performance high importance",
    "low": "Low performance low importance",
    "overkill": "High performance low importance",
}


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold: bool = False, color: str = TEXT, size: int = 11) -> Font:
    return Font(name="Calibri", bold=bold, color=color, size=size)


def _thin_border() -> Border:
    side = Side(style="thin", color=BLUE_MID)
    return Border(left=side, right=side, top=side, bottom=side)


def _safe_sheet_title(name: str, used: set[str]) -> str:
    cleaned = re.sub(r'[\\/*?:\[\]]', "-", str(name or "Sheet")).strip() or "Sheet"
    cleaned = cleaned[:31]
    base = cleaned
    n = 2
    while cleaned in used:
        suffix = f" ({n})"
        cleaned = (base[: 31 - len(suffix)] + suffix)[:31]
        n += 1
    used.add(cleaned)
    return cleaned


def _write_title(ws: Worksheet, title: str, subtitle: str | None = None) -> int:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    cell = ws.cell(row=1, column=1, value=title)
    cell.font = _font(bold=True, color=BLUE, size=16)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28
    row = 2
    if subtitle:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
        sub = ws.cell(row=2, column=1, value=subtitle)
        sub.font = _font(color=MUTED, size=10)
        sub.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[2].height = 20
        row = 3
    return row


def _style_header_row(ws: Worksheet, row: int, col_count: int) -> None:
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = _fill(BLUE)
        cell.font = _font(bold=True, color=WHITE, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _thin_border()


def _style_data_cell(
    cell,
    *,
    bold: bool = False,
    color: str | None = None,
    center: bool = False,
) -> None:
    cell.font = _font(bold=bold, color=color or TEXT)
    cell.alignment = Alignment(
        horizontal="center" if center else "left",
        vertical="center",
        wrap_text=True,
    )
    cell.border = _thin_border()


def _subtitle_bits(snapshot: dict[str, Any], filename: str = "") -> str:
    bits: list[str] = []
    if filename:
        bits.append(str(filename))
    elif snapshot.get("filename"):
        bits.append(str(snapshot["filename"]))
    if snapshot.get("scale"):
        bits.append(f"Scale {snapshot['scale']}")
    metric = snapshot.get("metric_label") or snapshot.get("metric")
    if metric:
        bits.append(f"Metric: {metric}")
    return " | ".join(bits)


def _quadrant_label(q: str | None) -> str:
    key = (q or "").strip().lower()
    return QUADRANT_LABELS.get(key, key or "-")


def _write_metric_block(
    ws: Worksheet,
    row: int,
    metrics: list[tuple[str, Any]],
) -> int:
    for label, value in metrics:
        ws.cell(row=row, column=1, value=label).font = _font(bold=True, color=MUTED, size=10)
        val_cell = ws.cell(row=row, column=2, value=value)
        val_cell.font = _font(bold=True, color=BLUE, size=11)
        row += 1
    return row + 1


def _write_statement_table(
    ws: Worksheet,
    row: int,
    headers: list[str],
    rows: list[list[Any]],
    *,
    col_widths: list[float] | None = None,
) -> int:
    col_count = len(headers)
    for col, text in enumerate(headers, start=1):
        ws.cell(row=row, column=col, value=text)
    _style_header_row(ws, row, col_count)
    row += 1

    for data_row in rows:
        for col, val in enumerate(data_row, start=1):
            cell = ws.cell(row=row, column=col, value=val)
            _style_data_cell(cell, center=col > 2)
        row += 1

    widths = col_widths or [42, 18, 22, 14, 14, 12, 12]
    for i, w in enumerate(widths[:col_count], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    return row


def build_bottom_up_xlsx(
    snapshot: dict[str, Any],
    result: dict[str, Any],
    *,
    filename: str = "",
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Bottom-up"
    subtitle = _subtitle_bits(snapshot, filename)
    row = _write_title(ws, "Simulation — Bottom-up", subtitle or None)

    row = _write_metric_block(
        ws,
        row,
        [
            ("Baseline overall satisfaction", f"{result.get('baseline_overall', '-')}%"),
            ("Predicted overall satisfaction", f"{result.get('predicted_overall', '-')}%"),
            ("Change (pts)", result.get("delta_pts", "-")),
            ("Method", result.get("method") or "sumproduct_reduced_importance"),
        ],
    )

    headers = [
        "Statement",
        "Section",
        "Quadrant",
        "Baseline performance %",
        "Simulated performance %",
        "Change (pts)",
        "Reduced importance %",
    ]
    table_rows: list[list[Any]] = []
    for stmt in result.get("statements") or []:
        base = stmt.get("performance")
        sim = stmt.get("simulated_performance", base)
        try:
            delta = round(float(sim) - float(base), 1)
        except (TypeError, ValueError):
            delta = "-"
        table_rows.append(
            [
                stmt.get("label") or stmt.get("column"),
                stmt.get("section_name") or stmt.get("section") or "-",
                _quadrant_label(stmt.get("quadrant")),
                base,
                sim,
                delta,
                stmt.get("reduced_importance"),
            ]
        )

    _write_statement_table(ws, row, headers, table_rows)
    ws.freeze_panes = "A5"
    ws.sheet_view.showGridLines = False

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_top_down_xlsx(
    snapshot: dict[str, Any],
    result: dict[str, Any],
    *,
    filename: str = "",
    sheet_title: str = "Top-down",
    page_title: str = "Simulation — Top-down",
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = _safe_sheet_title(sheet_title, set())
    subtitle = _subtitle_bits(snapshot, filename)
    row = _write_title(ws, page_title, subtitle or None)

    quadrants = result.get("quadrants") or []
    q_text = ", ".join(_quadrant_label(q) for q in quadrants) if quadrants else "None selected"

    metrics: list[tuple[str, Any]] = [
        ("Baseline overall satisfaction", f"{result.get('baseline_overall', '-')}%"),
        ("Target overall satisfaction", f"{result.get('target_overall', '-')}%"),
        ("Achieved overall satisfaction", f"{result.get('achieved_overall', '-')}%"),
        ("Selected quadrant(s)", q_text),
        ("Statements changed", result.get("changed_count", "-")),
    ]
    if result.get("note"):
        metrics.append(("Notes", result.get("note")))
    row = _write_metric_block(ws, row, metrics)

    option = result.get("option") or {}
    statements = option.get("statements") or result.get("changed_statements") or []

    headers = [
        "Statement",
        "Section",
        "Quadrant",
        "Before %",
        "After %",
        "Change (pts)",
        "Changed",
    ]
    table_rows: list[list[Any]] = []
    for stmt in statements:
        before = stmt.get("performance")
        after = stmt.get("required_performance", before)
        delta = stmt.get("delta_pts")
        if delta is None:
            try:
                delta = round(float(after) - float(before), 1)
            except (TypeError, ValueError):
                delta = "-"
        changed = stmt.get("changed")
        if changed is None:
            try:
                changed = float(after) != float(before)
            except (TypeError, ValueError):
                changed = False
        table_rows.append(
            [
                stmt.get("label") or stmt.get("column"),
                stmt.get("section_name") or stmt.get("section") or "-",
                _quadrant_label(stmt.get("quadrant")),
                before,
                after,
                delta,
                "Yes" if changed else "No",
            ]
        )

    if not table_rows:
        ws.cell(row=row, column=1, value="No statement scores changed for this run.").font = _font(
            color=MUTED, size=10
        )
    else:
        _write_statement_table(ws, row, headers, table_rows)

    ws.freeze_panes = "A5"
    ws.sheet_view.showGridLines = False

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_summary_xlsx(
    snapshot: dict[str, Any],
    summary: dict[str, Any],
    *,
    filename: str = "",
) -> bytes:
    td = summary.get("top_down_result") or {}
    if not td:
        raise ValueError("No top-down simulation saved. Run simulation first.")
    merged = {
        **td,
        "quadrants": td.get("quadrants") or summary.get("top_down_quadrants") or [],
        "changed_statements": td.get("changed_statements") or [],
    }
    return build_top_down_xlsx(
        snapshot,
        merged,
        filename=filename,
        sheet_title="Summary",
        page_title="Summary — Top-down simulation",
    )
